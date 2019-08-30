from sqlalchemy import or_

from common.models import Session, ProductModel, ReviewModel, ImageModel
from settings import EXCEL_WB_PATH
from openpyxl import Workbook, styles


def write_model_to_worksheet(wb, model, session):
	ws = wb.create_sheet(model.__tablename__)
	ws.append(model.__table__.columns.keys())
	for product in session.query(model).all():
		data = list(map(product.__dict__.get, model.__table__.columns.keys()))
		ws.append(data)


def write_model_to_worksheet_all(wb, model, session):
	ws = wb.create_sheet(model.__tablename__)
	header = model.__table__.columns.keys()
	header += ['weight', 'price_for_100_original', 'price_for_100_final']
	ws.append(header)
	for product in session.query(model).all():
		data = list(map(product.__dict__.get, model.__table__.columns.keys()))

		data = transform_field(data, model.__table__.columns.keys(), mutations)

		data += [product.weight, product.price_for_100_original, product.price_for_100_final]
		ws.append(data)


def write_model_to_worksheet_filtered(wb, model, session):
	ws = wb.create_sheet(model.__tablename__)
	header = model.__table__.columns.keys()
	header += ['weight', 'price_for_100_original', 'price_for_100_final']
	ws.append(header)
	conditions = []
	for name in ['пшен', 'полб', 'рожь', 'ячме']:
		conditions.append(ProductModel.specs.ilike(f'%{name}%'))

	q = session.query(ProductModel).filter(
		~or_(*conditions)
	)

	for product in q.all():
		data = list(map(product.__dict__.get, model.__table__.columns.keys()))

		# transform img_url
		data = transform_field(data, model.__table__.columns.keys(), mutations)

		data += [product.weight, product.price_for_100_original, product.price_for_100_final]
		ws.append(data)


def transform_img_url(url):
	return f'http:{url}'


def transform_img_local_path(path):
	from os.path import sep
	return path.split(sep)[-1]


def transform_field(data, columns, field_func_dict):
	for k, v in field_func_dict.items():
		index = columns.index(k)
		data[index] = v(data[index])
	return data


def write_model_to_worksheet_highlight(wb, model, session):
	red_col = styles.colors.Color(rgb='00FF0000')
	red_fill = styles.fills.PatternFill(patternType='solid', fgColor=red_col)
	ws = wb.create_sheet(model.__tablename__)
	header = model.__table__.columns.keys()
	header += ['weight', 'price_for_100_original', 'price_for_100_final']
	ws.append(header)
	# for product in session.query(model).all():
	for product in session.query(model).all():
		data = list(map(product.__dict__.get, model.__table__.columns.keys()))

		#transform img_url
		data = transform_field(data, model.__table__.columns.keys(), mutations)

		data += [product.weight, product.price_for_100_original, product.price_for_100_final]
		ws.append(data)
		specs = product.specs or ''
		if any(i in specs for i in ['пшен', 'полб', 'рожь', 'ячме']):
			ws.row_dimensions[ws._current_row].fill = red_fill


def dump_to_excel(session, func, suffix='', write_only=True):
	wb = Workbook(write_only=write_only)
	func(wb, ProductModel, session)
	write_model_to_worksheet(wb, ReviewModel, session)
	write_model_to_worksheet(wb, ImageModel, session)
	wb.save(EXCEL_WB_PATH(suffix))


if __name__ == '__main__':
	session = Session()
	# dump_to_excel(session, write_model_to_worksheet_all, '_all')
	mutations = {
		# 'img_url': transform_img_url,
		# 'img_local_path': transform_img_local_path,
	}
	# dump_to_excel(session, write_model_to_worksheet_filtered, '_filtered')
	dump_to_excel(session, write_model_to_worksheet_highlight, '_highlight', write_only=False)
