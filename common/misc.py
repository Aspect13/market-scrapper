import json
from json import JSONDecodeError
from pathlib import Path
import functools

from common.logger_custom import logger
from settings import CACHE_DICT_PATH, LINK_SET_PATH


def pathify(folder_path):
	def decorated_path(func):
		@functools.wraps(func)
		def wrapper(*args, **kwargs):
			try:
				file_name = func(*args, **kwargs)
			except KeyError:
				print(f'No cache for url: {args[0]}')
				return None
			return Path.joinpath(folder_path, file_name)
		return wrapper
	return decorated_path


def get_cache_dict():
	try:
		return json.load(open(CACHE_DICT_PATH, 'r'))
	except FileNotFoundError:
		with open(CACHE_DICT_PATH, 'w') as f:
			f.write('{}')
	return {}


def get_link_set(force_update_from_xlsx=False):
	link_set = []
	try:
		link_set = json.load(open(LINK_SET_PATH, 'r', encoding='utf8'))
		assert link_set
		assert not force_update_from_xlsx
	except (FileNotFoundError, AssertionError, JSONDecodeError):
		try:
			from openpyxl import load_workbook
			from settings import IO_PATH
			wb = load_workbook(Path(IO_PATH, 'link_set.xlsx'))
			ws = wb.active
			for row in ws.iter_rows(min_row=2):
				item = {
					'category_name': row[0].value,
					'url': row[1].value
				}

				# if not any(item['url'] == i.url for i in link_set):
				if (not force_update_from_xlsx) or (force_update_from_xlsx and item not in link_set):
					link_set.append(item)

			json.dump(link_set, open(LINK_SET_PATH, 'w', encoding='utf8'))
		except FileNotFoundError as e:
			logger.critical('No link set provided')
			raise e

	return link_set


def add_schema(url, schema='http'):
	if not url.startswith(schema):
		return f'{schema}:{url}'
	return url
